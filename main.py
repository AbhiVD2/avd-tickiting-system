from datetime import datetime, timedelta

import pandas as pd
from flask import Flask, send_file
from flask_cors import CORS
from flask_pymongo import PyMongo
from flask_restx import Api
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# MongoDB setup (replace the URI if needed)
app.config["MONGO_URI"] = "mongodb://localhost:27017/ticket_system"
mongo = PyMongo(app)

# Collection references
societyData = mongo.db.societyData
managers = mongo.db.managers
deleted_managers = mongo.db.deleted_managers
users_collection = mongo.db.memberMaster
deleted_users_collection = mongo.db.deleted_users
ticketdata = mongo.db.tickets
members_collection = mongo.db.memberMaster
events_collection = mongo.db.events_collection
# Initialize Flask-RESTPlus API
api = Api(app, version='1.0', title='Ticket System API', description='A simple API for managing ticket systems')


# societies_collection = mongo.db.societiesCollection

# Helper function to convert ObjectId to string
def objectid_to_str(record):
    if '_id' in record:
        record['_id'] = str(record['_id'])
    return record


# Societies Routes

from bson import ObjectId
from flask import jsonify, request

@app.route('/api/societies', methods=['GET'])
def get_societies():
    member_id = request.args.get("member_id")
    manager_id = request.args.get("manager_id")

    if member_id:
        try:
            member = members_collection.find_one({'_id': ObjectId(member_id)})
            if member:
                societies = societyData.find({'members': member_id})
                society_list = []
                
                for society in societies:
                    # Convert ObjectId fields to string (e.g. '_id' field)
                    society['_id'] = str(society['_id'])
                    society_list.append(society)

                return jsonify(society_list)
            else:
                return jsonify({"message": "No society found for provided member"}), 404
        except Exception as e:
            return jsonify({"message": "Database error", "error": str(e)}), 500
    if manager_id:
        try:
            member = managers.find_one({'_id': ObjectId(manager_id)})
            if member:
                societies = societyData.find({'managers': member_id})
                society_list = []

                for society in societies:
                    # Convert ObjectId fields to string (e.g. '_id' field)
                    society['_id'] = str(society['_id'])
                    society_list.append(society)

                return jsonify(society_list)
            else:
                return jsonify({"message": "No society found for provided member"}), 404
        except Exception as e:
            return jsonify({"message": "Database error", "error": str(e)}), 500
    else:
        try:
            all_societies = list(societyData.find())
            for society in all_societies:
                # Convert ObjectId fields to string (e.g. '_id' field)
                society['_id'] = str(society['_id'])
            return jsonify(all_societies)
        except Exception as e:
            return jsonify({"message": "Database error", "error": str(e)}), 500

        

@app.route('/api/societies', methods=['POST'])
def add_society():
    data = request.json
    new_society = {
        "name": data.get('name'),
        "address": data.get('address'),
        "incharge": data.get('incharge'),
        "contact": data.get('contact')
    }
    result = societyData.insert_one(new_society)
    return jsonify({"message": "Society added", "id": str(result.inserted_id)}), 201


@app.route('/api/societies/<society_id>', methods=['PUT'])
def update_society(society_id):
    data = request.json
    # updated_data = {
    #     "name": data.get('name'),
    #     "address": data.get('address'),
    #     "incharge": data.get('incharge'),
    #     "contact": data.get('contact')
    # }
    societyData.update_one({'_id': ObjectId(society_id)}, {'$set': data})
    return jsonify({"message": "Society updated successfully"})


@app.route('/api/societies/<society_id>', methods=['DELETE'])
def delete_society(society_id):
    societyData.delete_one({'_id': ObjectId(society_id)})
    return jsonify({"message": "Society deleted successfully"})


# Manager Routes

@app.route('/api/managers', methods=['GET'])
def get_managers():
    all_managers = list(managers.find()) 

    for manager in all_managers:
        manager['_id'] = str(manager['_id'])
    return jsonify(all_managers)


@app.route('/api/managers', methods=['POST'])
def add_manager():
    data = request.json
    result = managers.insert_one(data)
    return jsonify({'message': 'Manager added successfully', 'id': str(result.inserted_id)}), 201


@app.route('/api/managers/<manager_id>', methods=['PUT'])
def update_manager(manager_id):
    updated_manager = request.json
    result = managers.update_one({'_id': ObjectId(manager_id)}, {'$set': updated_manager})
    if result.matched_count:
        return jsonify({"message": "Manager updated successfully!"})
    return jsonify({"message": "Manager not found!"}), 404


@app.route('/api/managers/<manager_id>', methods=['DELETE'])
def delete_manager(manager_id):
    manager = managers.find_one({'_id': ObjectId(manager_id)})
    if manager:
        manager['deletedAt'] = datetime.utcnow()
        deleted_managers.insert_one(manager)
        managers.delete_one({'_id': ObjectId(manager_id)})
        return jsonify({"message": "Manager deleted successfully!"})
    return jsonify({"message": "Manager not found!"}), 404


@app.route('/api/managers/restore/<manager_id>', methods=['POST'])
def restore_manager(manager_id):
    manager = deleted_managers.find_one({'_id': ObjectId(manager_id)})
    if manager:
        del manager['deletedAt']
        managers.insert_one(manager)
        deleted_managers.delete_one({'_id': ObjectId(manager_id)})
        return jsonify({"message": "Manager restored successfully!"})
    return jsonify({"message": "Manager not found!"}), 404


@app.route('/api/managers/permanent_delete/<manager_id>', methods=['DELETE'])
def permanently_delete_manager(manager_id):
    result = deleted_managers.delete_one({'_id': ObjectId(manager_id)})
    if result.deleted_count:
        return jsonify({"message": "Manager permanently deleted!"})
    return jsonify({"message": "Manager not found!"}), 404


@app.route('/api/deleted_managers', methods=['GET'])
def get_deleted_managers():
    one_month_ago = datetime.utcnow() - timedelta(days=30)
    deleted_managers_list = list(deleted_managers.find({'deletedAt': {'$gt': one_month_ago}}))
    return jsonify(deleted_managers_list)


# Users Routes

@app.route('/api/users', methods=['GET'])
def get_users():
    active_users = [objectid_to_str(user) for user in users_collection.find({"deletedAt": {"$exists": False}})]
    return jsonify(active_users)


@app.route('/api/users', methods=['POST'])
def update_user():
    data = request.json
    user_id = data.get('_id')
    updated_data = {key: value for key, value in data.items() if key != '_id'}

    result = users_collection.update_one({'_id': ObjectId(user_id)}, {'$set': updated_data})

    if result.matched_count:
        return jsonify({"message": "User updated successfully!"})
    return jsonify({"message": "User not found!"}), 404


@app.route('/api/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    user = users_collection.find_one({'_id': ObjectId(user_id)})
    if user:
        user['deletedAt'] = datetime.utcnow()
        deleted_users_collection.insert_one(user)
        users_collection.delete_one({'_id': ObjectId(user_id)})
        return jsonify({"message": "User soft deleted successfully!"})
    return jsonify({"message": "User not found!"}), 404


@app.route('/api/users/restore/<user_id>', methods=['POST'])
def restore_user(user_id):
    deleted_user = deleted_users_collection.find_one({'_id': ObjectId(user_id)})
    if deleted_user:
        del deleted_user['deletedAt']
        users_collection.insert_one(deleted_user)
        deleted_users_collection.delete_one({'_id': ObjectId(user_id)})
        return jsonify({"message": "User restored successfully!"})
    return jsonify({"message": "User not found!"}), 404


@app.route('/api/users/permanent/<user_id>', methods=['DELETE'])
def permanently_delete_user(user_id):
    result = deleted_users_collection.delete_one({'_id': ObjectId(user_id)})
    if result.deleted_count:
        return jsonify({"message": "User permanently deleted!"})
    return jsonify({"message": "User not found!"}), 404


@app.route('/api/deleted-users', methods=['GET'])
def get_deleted_users():
    deleted_users = [objectid_to_str(user) for user in deleted_users_collection.find()]
    return jsonify(deleted_users)


@app.route('/api/deleted-users/cleanup', methods=['DELETE'])
def clean_up_deleted_users():
    one_month_ago = datetime.utcnow() - timedelta(days=30)
    deleted_users_collection.delete_many({'deletedAt': {'$lt': one_month_ago.isoformat()}})
    return jsonify({"message": "Deleted users older than one month have been removed."})


@app.route('/api/deleted_managers/<manager_id>/restore', methods=['PUT'])
def api_restore_manager(manager_id):
    try:
        # Find the manager in the deleted managers collection
        manager = deleted_managers.find_one({'_id': ObjectId(manager_id)})

        if not manager:
            return jsonify({"message": "Manager not found!"}), 404

        # Remove the 'deletedAt' field to restore the manager
        del manager['deletedAt']

        # Insert the manager back into the managers collection
        managers.insert_one(manager)

        # Remove the manager from the deleted managers collection
        deleted_managers.delete_one({'_id': ObjectId(manager_id)})

        return jsonify({"message": "Manager restored successfully!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


#

@app.route('/api/tickets', methods=['GET'])
def get_tickets():
    # Get manager_id from query parameters (optional)
    manager_id = request.args.get('manager_id')

    # If manager_id is provided, filter by that manager_id
    if manager_id:
        tickets = ticketdata.find({"assigned_to": manager_id})
    else:
        # If no manager_id is provided, fetch all tickets
        tickets = ticketdata.find()

    # Prepare the ticket data to be returned
    ticket_list = []
    for ticket in tickets:
        ticket['_id'] = str(ticket['_id'])  # Convert ObjectId to string
        ticket_list.append(ticket)

    return jsonify(ticket_list)


@app.route('/api/tickets', methods=['POST'])
def add_ticket():
    ticket = request.json

    # Assuming you have a way to get the logged-in user's manager ID
    logged_in_user = request.headers.get('X-Logged-In-User')  # Example header for logged-in user
    if logged_in_user:
        # Fetch the manager ID for the logged-in user
        manager = managers.find_one({"_id": ObjectId(logged_in_user)})
        if manager:
            ticket['manager_id'] = str(manager['_id'])  # Assign the manager ID to the ticket
        else:
            return jsonify({"message": "Manager not found for the logged-in user."}), 404
    else:
        # If no user is logged in, assign the ticket to the manager associated with the society
        society_id = ticket.get('society_id')  # Assuming the society ID is sent with the ticket data
        if society_id:
            society = mongo.db.societyData.find_one({"_id": ObjectId(society_id)})
            if society and 'managers' in society:
                ticket['manager_id'] = society['managers']  # Assign the first manager associated with the society
            else:
                return jsonify({"message": "No manager found for the selected society."}), 404
        else:
            return jsonify({"message": "Society ID is required."}), 400

    # Insert the ticket into the database
    ticketdata.insert_one(ticket)
    return jsonify({"message": "Ticket added successfully!"}), 201


@app.route('/api/tickets/<ticket_id>', methods=['PUT'])
def update_ticket(ticket_id):
    updated_ticket = request.json
    ticketdata.update_one({'_id': ObjectId(ticket_id)}, {'$set': updated_ticket})
    return jsonify({"message": "Ticket updated successfully!"})


@app.route('/api/tickets/<ticket_id>', methods=['DELETE'])
def delete_ticket(ticket_id):
    ticketdata.delete_one({'_id': ObjectId(ticket_id)})
    return jsonify({"message": "Ticket deleted successfully!"})


# uploading excel bulk staff
@app.route('/upload', methods=['POST'])
def upload_file():
    # Check if the file is part of the request
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    # Extract the key (Society ID) from the request parameters
    key = request.args.get('society_id')

    if not key:
        return jsonify({"error": "No key provided"}), 400

    # Check if the file is an Excel file
    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "File is not an Excel file"}), 400

    # Read the Excel file into a DataFrame
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Column mapping
    column_mapping = {
        "Member Type": "member_type",
        "Flat Owner Name": "flat_owner_name",
        "Tenant Name": "tenant_name",
        "createUsername": "createUsername",
        "createPassword": "createPassword",
        "Wing/Building No": "building_number",
        "Flat No": "flat_number",
        "Flat Type(1,2,3 Bhk)": "flat_type",
        "Primary Contact No": "primary_contact_number",
        "Email Id": "email",
        "Owner Contact No": "owner_contact_number",
        "Tenant Contact No": "tenant_contact_number",
        "2 Wheelers": "two_wheelers",
        "4 Wheeler": "four_wheelers",
        "Stilt": "stilt",
        "Parking Charges": "parking_charges",
        "Maintenance Charges": "maintenance_charges",
        "Outstanding": "outstanding",
        "Maid": "maid",
        "Pet": "pet",
        "Share Certificate": "share_certificate",
        "Remark": "remark"
    }

    # Get the expected columns based on the mapping keys
    expected_columns = list(column_mapping.keys())

    # Check if all expected columns are present in the uploaded Excel file
    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        return jsonify({"error": f"Missing columns in the Excel file: {', '.join(missing_columns)}"}), 400

    # Transform DataFrame to mapped data
    data = []
    for _, row in df[expected_columns].iterrows():
        # Transform each row using the column mapping
        transformed_record = {
            column_mapping.get(incoming_column): row[incoming_column]
            for incoming_column in expected_columns
        }
        transformed_record['society_id'] = key
        data.append(transformed_record)

    # Insert data into Members Collection
    try:
        if data:
            # Insert members into the Members collection
            inserted_members = members_collection.insert_many(data)

            # Optionally, update the Society collection to include the new member references
            society = societyData.find_one({"_id": ObjectId(key)})
            updated_member_ids = [str(member_id) for member_id in inserted_members.inserted_ids]

            if society:
                societyData.update_one(
                    {"_id": ObjectId(key)},
                    {"$push": {"members": {"$each": updated_member_ids}}}
                )
            else:
                return jsonify({"error": "Society not found"}), 400

            return jsonify({"message": "Data inserted successfully", "count": len(inserted_members.inserted_ids)}), 201
        else:
            return jsonify({"message": "No data to insert"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

import tempfile

@app.route('/download', methods=['GET'])
def download_excel():
    # Sample data (this could be fetched from a database or other source)
    data = [
        {"Member Type": "Owner/SCT", "Flat Owner Name": "John Doe", "Tenant Name": "Jane Smith", "Flat No": "101",
         "Wing/Building No": "A", "Primary Contact No": "1234567890", "Email Id": "john.doe@example.com",
         "Owner Contact No": "1231231234", "Tenant Contact No": "9879879876", "2 Wheelers": 1, "4 Wheeler": 1,
         "Stilt": 1, "Parking Charges": 500, "Maintenance Charges": 300, "Outstanding": 100, "Maid": "Yes",
         "Pet": "No", "Share Certificate": "Yes", "Remark": "No remark"}
    ]

    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Use a NamedTemporaryFile to avoid security issues
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        excel_path = tmp_file.name  # Get the temp file path

    # Create an Excel file and save data
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Members Data")

    # Load the workbook to modify it
    wb = load_workbook(excel_path)
    sheet = wb["Members Data"]

    # Add columns for createUsername and createPassword at the end
    create_username_col = len(df.columns) + 1
    create_password_col = len(df.columns) + 2

    # Set headers for new columns
    sheet.cell(row=1, column=create_username_col, value="createUsername")
    sheet.cell(row=1, column=create_password_col, value="createPassword")

    # Add formulas
    for row in range(2, sheet.max_row + 1):  # Start from row 2 (avoid header)
        email_cell = sheet.cell(row=row, column=7)  # "Email Id" in column G (7)
        owner_name_cell = sheet.cell(row=row, column=2)  # "Flat Owner Name" in column B (2)
        primary_contact_cell = sheet.cell(row=row, column=9)  # "Primary Contact No" in column I (9)

        # Formula for createUsername
        sheet.cell(row=row, column=create_username_col, value=f"={email_cell.coordinate}")

        # Formula for createPassword
        sheet.cell(row=row, column=create_password_col,
                   value=f"=LEFT({owner_name_cell.coordinate},4)&\"@\"&RIGHT({primary_contact_cell.coordinate},4)")

    # Save modified workbook
    wb.save(excel_path)

    # Send the file as an attachment
    return send_file(excel_path, as_attachment=True, download_name="members_data_with_formulas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.route('/get_members/<society_id>', methods=['GET'])
def get_members(society_id):
    try:
        # Retrieve all members for a given society
        members = list(members_collection.find({"society_id": society_id}))
        return jsonify(members), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def parse_object_id(obj_id):
    try:
        return ObjectId(obj_id)
    except Exception as e:
        return None


# CREATE event
@app.route('/api/events', methods=['POST'])
def create_event():
    data = request.get_json()

    if not data.get('title') or not data.get('date') or not data.get('time'):
        return jsonify({"error": "Missing required fields: title, date, or time"}), 400

    event_data = {
        "title": data['title'],
        "date": data['date'],  # Store date as string in 'YYYY-MM-DD' format
        "time": data['time'],  # Store time as string in 'HH:MM' format
        "location": data.get('location', ''),
        "created_at": datetime.utcnow()
    }
    result = events_collection.insert_one(event_data)

    return jsonify({"message": "Event created successfully", "event_id": str(result.inserted_id)}), 201

# GET all events
@app.route('/api/events', methods=['GET'])
def get_all_events():
    events = events_collection.find()
    events_list = []
    for event in events:
        event['_id'] = str(event['_id'])  # Convert ObjectId to string
        events_list.append(event)
    return jsonify(events_list), 200

# GET event by ID
@app.route('/api/events/<event_id>', methods=['GET'])
def get_event(event_id):
    event = events_collection.find_one({"_id": parse_object_id(event_id)})
    if not event:
        return jsonify({"error": "Event not found"}), 404
    event['_id'] = str(event['_id'])
    return jsonify(event), 200

# UPDATE event by ID
@app.route('/api/events/<event_id>', methods=['PUT'])
def update_event(event_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    update_data = {}
    if 'title' in data:
        update_data['title'] = data['title']
    if 'date' in data:
        update_data['date'] = data['date']
    if 'time' in data:
        update_data['time'] = data['time']
    if 'location' in data:
        update_data['location'] = data['location']

    result = events_collection.update_one(
        {"_id": parse_object_id(event_id)},
        {"$set": update_data}
    )

    if result.modified_count == 0:
        return jsonify({"error": "No changes made"}), 400

    return jsonify({"message": "Event updated successfully", "event_id": event_id}), 200

# DELETE event by ID
@app.route('/api/events/<event_id>', methods=['DELETE'])
def delete_event(event_id):
    event = events_collection.find_one({"_id": parse_object_id(event_id)})
    if not event:
        return jsonify({"error": "Event not found"}), 404

    events_collection.delete_one({"_id": parse_object_id(event_id)})
    return jsonify({"message": "Event deleted successfully", "event_id": event_id}), 200

if __name__ == '__main__':
    app.run(debug=True)
